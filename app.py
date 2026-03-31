"""
app.py — Descarga Automática Certificados de Tope SOAT
Flask + Playwright · Activa IT / RGC
"""

import asyncio
import base64
import json
import io
import os
from pathlib import Path
from datetime import datetime
from queue import Queue
# Now you can create a Queue object like this:
import pandas as pd
from flask import Flask, render_template, request, jsonify, Response, send_file
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
queue = Queue(maxsize=0)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

# ─── ESTADO GLOBAL ────────────────────────────────────────────────────────────
# Una sola tarea a la vez (uso local)
estado = {
    "corriendo":   False,
    "progreso":    [],      # lista de dicts {fila, tipo_doc, numero, estado, archivo, error, ...}
    "log_queue":   queue.Queue(),
    "total":       0,
    "terminado":   False,
    "pdfs":        {},      # {nombre: bytes}
    "config":      {},
}

# ─── MAPEOS ───────────────────────────────────────────────────────────────────
TIPO_DOC_TEXTO = {
    "CC":"Cedula Ciudadanía","CE":"Cedula Extranjería","TI":"Tarjeta de Identidad",
    "RC":"Registro Civil","MS":"Menor Sin Identificación","AS":"Adulto Sin Identificación",
    "PA":"Pasaporte","DE":"Documento Extranjero","NI":"NIT",
    "CN":"Certificado Nacido Vivo","CDF":"Certificado de Defunción",
}
TIPO_AMPARO_TEXTO = {
    "MED":"Gastos Médicos - Quirúrgicos, Farmacéuticos y Hospitalarios",
    "TRA":"Indemnización por Gastos de Transporte y Movilización de Víctimas",
    "FUN":"Indemnización por Gastos Funerarios",
    "PER":"Indemnización por Incapacidad Permanente",
    "MUE":"Indemnización por muerte",
}

# ─── LÓGICA PLAYWRIGHT (asíncrona) ────────────────────────────────────────────

def log(msg: str):
    estado["log_queue"].put({"tipo": "log", "msg": msg})

def progreso_row(row: dict):
    estado["log_queue"].put({"tipo": "row", "data": row})


async def login(page, usuario, password, login_url):
    log("🔐 Iniciando sesión en Activa IT...")
    await page.goto(login_url, wait_until="networkidle")
    await page.fill('input[placeholder="Usuario"]', usuario)
    await page.fill('input[placeholder="Contraseña"]', password)
    checkbox = page.locator('input[type="checkbox"]')
    if not await checkbox.is_checked():
        await checkbox.check()
    await page.click('button:has-text("Inicio de sesión"), input[value="Inicio de sesión"]')
    await page.wait_for_url("**/Index.aspx", timeout=15000)
    log("✅ Sesión iniciada correctamente")


async def navegar_certificado_tope(page):
    await page.click('text=Tareas Auditoria')
    await page.wait_for_timeout(800)
    await page.click('text=Certificado de Tope')
    await page.wait_for_timeout(2500)
    try:
        aceptar = page.locator('button:has-text("Aceptar"), input[value="Aceptar"]')
        if await aceptar.is_visible(timeout=3000):
            await aceptar.click()
            await page.wait_for_timeout(800)
    except:
        pass
    log("📋 Formulario Certificado de Tope cargado")


async def get_frame(page):
    for f in page.frames:
        if "CertificadoTope" in f.url:
            return f
    raise Exception("No se encontró el iframe CertificadoTope")


async def llenar_formulario(frame, tipo_doc, numero, tipo_amparo):
    txt_doc = TIPO_DOC_TEXTO.get(tipo_doc, tipo_doc).replace("'", "\\'")
    txt_amp = TIPO_AMPARO_TEXTO.get(tipo_amparo, tipo_amparo).replace("'", "\\'")
    await frame.evaluate(f"""
        (() => {{
            const selDoc = document.querySelector('#ddlTipoDoc_cmbParametros');
            const txtDoc = document.querySelector('#rgctxtddlTipoDoc_cmbParametros');
            if (selDoc) selDoc.value = '{tipo_doc}';
            if (txtDoc) txtDoc.value = '{txt_doc}';
            ['change','input','blur'].forEach(ev => {{
                if (selDoc) selDoc.dispatchEvent(new Event(ev, {{bubbles:true}}));
                if (txtDoc) txtDoc.dispatchEvent(new Event(ev, {{bubbles:true}}));
            }});
            const tb = document.querySelector('#tbNumDoc');
            if (tb) {{ tb.value = '{numero}'; tb.dispatchEvent(new Event('change', {{bubbles:true}})); }}
            const selAmp = document.querySelector('#ddlTipo_cmbParametros');
            const txtAmp = document.querySelector('#rgctxtddlTipo_cmbParametros');
            if (selAmp) selAmp.value = '{tipo_amparo}';
            if (txtAmp) txtAmp.value = '{txt_amp}';
            ['change','input','blur'].forEach(ev => {{
                if (selAmp) selAmp.dispatchEvent(new Event(ev, {{bubbles:true}}));
                if (txtAmp) txtAmp.dispatchEvent(new Event(ev, {{bubbles:true}}));
            }});
        }})();
    """)
    await frame.wait_for_timeout(300)


async def buscar(frame):
    await frame.evaluate("document.querySelector('#btSearch').click();")
    await frame.wait_for_timeout(3500)


async def hay_resultados(frame) -> bool:
    try:
        filas = await frame.query_selector_all('#gvSiniestros tr')
        return len(filas) >= 2
    except:
        return False


async def obtener_info_siniestro(frame) -> dict:
    info = {"nombre_victima": "", "poliza": "", "siniestro": "", "placa": ""}
    try:
        body = await frame.inner_text('body')
        m = re.search(r'Nombre víctima:\s*(.+)', body)
        if m:
            info["nombre_victima"] = m.group(1).strip()
        filas = await frame.query_selector_all('#gvSiniestros tr')
        if len(filas) >= 2:
            celdas = await filas[1].query_selector_all('td')
            vals = [(await c.inner_text()).strip() for c in celdas]
            if len(vals) >= 1: info["poliza"]    = vals[0]
            if len(vals) >= 2: info["siniestro"] = vals[1]
            if len(vals) >= 3: info["placa"]     = vals[2]
    except:
        pass
    return info


async def descargar_pdf(page, frame, idx: int) -> dict:
    from playwright.async_api import TimeoutError as PlaywrightTimeout
    result = {"estado": "SIN_PDF", "archivo": "", "error": "", "pdf_bytes": None}
    try:
        async with page.expect_response(
            lambda r: "CrearCertificado" in r.url, timeout=15000
        ) as resp_info:
            clicked = await frame.evaluate("""
                (() => {
                    const rows = document.querySelectorAll('#gvSiniestros tr');
                    if (rows.length < 2) return 'NO_ROWS';
                    const btn = rows[1].querySelector('input[type="image"]');
                    if (btn) { btn.click(); return 'OK'; }
                    const pdfBtn = rows[1].querySelector('input[src*="PDF"],input[src*="pdf"]');
                    if (pdfBtn) { pdfBtn.click(); return 'OK_PDF_SRC'; }
                    const lastCell = rows[1].querySelector('td:last-child');
                    if (lastCell) {
                        const any = lastCell.querySelector('*');
                        if (any) { any.click(); return 'OK_LAST'; }
                        lastCell.click(); return 'OK_LAST_CELL';
                    }
                    return 'NOT_FOUND';
                })();
            """)

        response  = await resp_info.value
        body_text = await response.text()
        outer     = json.loads(body_text)
        inner     = json.loads(outer.get("d", "{}"))

        errores = inner.get("Errores", "").strip()
        b64     = inner.get("Ruta", "").strip()

        if errores:
            result["estado"] = "ERROR_SERVIDOR"
            result["error"]  = errores
        elif not b64:
            result["estado"] = "SIN_BASE64"
            result["error"]  = "Respuesta sin PDF"
        else:
            pdf_bytes = base64.b64decode(b64)
            nombre    = f"cert_tope_{idx:03d}.pdf"
            result["estado"]    = "DESCARGADO"
            result["archivo"]   = nombre
            result["pdf_bytes"] = pdf_bytes

    except PlaywrightTimeout:
        result["estado"] = "TIMEOUT"
        result["error"]  = "Sin respuesta del servidor (15s)"
    except Exception as e:
        result["estado"] = "ERROR"
        result["error"]  = str(e)

    return result


async def run_playwright(cfg: dict, registros: list):
    from playwright.async_api import async_playwright
    import asyncio

    for idx, row in enumerate(registros, 1):
        tipo_doc    = str(row.get("TIPO", "")).strip().upper()
        numero      = str(row.get("ID", "")).strip()
        tipo_amparo = str(row.get("TIPO DE AMPARO", "")).strip().upper()

        resultado = {
            "fila": idx, "tipo_doc": tipo_doc, "numero": numero,
            "tipo_amparo": tipo_amparo, "nombre_victima": "",
            "poliza": "", "siniestro": "", "placa": "",
            "estado": "", "archivo": "", "error": ""
        }

        log(f"[{idx}/{len(registros)}] {tipo_doc} {numero} — {tipo_amparo}")

        # Solo disponible cuando browser está activo (se pasa por closure)
        resultado["_pending"] = True
        estado["progreso"].append(resultado)
        progreso_row({"idx": idx - 1, "data": resultado, "total": len(registros)})

    # Reset y correr de verdad
    estado["progreso"].clear()

    async with async_playwright() as p:
        headless = cfg.get("headless", False)
        browser  = await p.chromium.launch(headless=headless)
        context  = await browser.new_context(viewport={"width": 1280, "height": 800})
        page     = await context.new_page()

        try:
            await login(page, cfg["usuario"], cfg["password"], cfg["login_url"])
            await navegar_certificado_tope(page)
            frame = await get_frame(page)

            for idx, row in enumerate(registros, 1):
                tipo_doc    = str(row.get("TIPO", "")).strip().upper()
                numero      = str(row.get("ID", "")).strip()
                tipo_amparo = str(row.get("TIPO DE AMPARO", "")).strip().upper()

                resultado = {
                    "fila": idx, "tipo_doc": tipo_doc, "numero": numero,
                    "tipo_amparo": tipo_amparo, "nombre_victima": "",
                    "poliza": "", "siniestro": "", "placa": "",
                    "estado": "PROCESANDO", "archivo": "", "error": ""
                }

                log(f"[{idx}/{len(registros)}] Buscando {tipo_doc} {numero}…")
                progreso_row({"idx": idx - 1, "data": resultado, "total": len(registros)})

                try:
                    await llenar_formulario(frame, tipo_doc, numero, tipo_amparo)
                    await buscar(frame)

                    if not await hay_resultados(frame):
                        resultado["estado"] = "NO_ENCONTRADO"
                        log(f"  ⚠️  Sin resultados para {tipo_doc} {numero}")
                    else:
                        info = await obtener_info_siniestro(frame)
                        resultado.update(info)
                        r = await descargar_pdf(page, frame, idx)
                        pdf_bytes = r.pop("pdf_bytes", None)
                        resultado.update(r)
                        if pdf_bytes:
                            estado["pdfs"][r["archivo"]] = pdf_bytes
                            log(f"  ✅ PDF descargado: {r['archivo']} ({len(pdf_bytes):,} bytes)")
                        else:
                            log(f"  ❌ {r['estado']}: {r['error']}")

                except Exception as e:
                    resultado["estado"] = "ERROR"
                    resultado["error"]  = str(e)
                    log(f"  ❌ Error: {e}")

                estado["progreso"].append(resultado)
                progreso_row({"idx": idx - 1, "data": resultado, "total": len(registros)})
                await asyncio.sleep(cfg.get("delay", 2))

        except Exception as e:
            log(f"❌ Error crítico: {e}")
        finally:
            await browser.close()

    estado["corriendo"] = False
    estado["terminado"] = True
    log("🏁 Proceso completado")
    estado["log_queue"].put({"tipo": "fin"})


def hilo_playwright(cfg, registros):
    asyncio.run(run_playwright(cfg, registros))


# ─── EXCEL ────────────────────────────────────────────────────────────────────

def generar_excel(registros: list) -> bytes:
    df = pd.DataFrame(registros)
    cols_orden = ["fila","tipo_doc","numero","tipo_amparo","nombre_victima",
                  "poliza","siniestro","placa","estado","archivo","error"]
    df = df[[c for c in cols_orden if c in df.columns]]

    etiquetas = {
        "fila":"#","tipo_doc":"Tipo Doc","numero":"Nº Documento",
        "tipo_amparo":"Amparo","nombre_victima":"Nombre Víctima",
        "poliza":"Póliza","siniestro":"Siniestro","placa":"Placa",
        "estado":"Estado","archivo":"Archivo PDF","error":"Error",
    }
    df.rename(columns=etiquetas, inplace=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Descargas")
        ws = writer.sheets["Descargas"]

        navy   = PatternFill("solid", fgColor="0f1e35")
        verde  = PatternFill("solid", fgColor="27ae60")
        rojo   = PatternFill("solid", fgColor="c0392b")
        amber  = PatternFill("solid", fgColor="e67e22")
        gris   = PatternFill("solid", fgColor="ecf0f1")
        blanco = PatternFill("solid", fgColor="FFFFFF")
        thin   = Side(style='thin', color="bdc3c7")
        brd    = Border(left=thin,right=thin,top=thin,bottom=thin)
        c      = Alignment(horizontal="center",vertical="center")
        iz     = Alignment(horizontal="left",  vertical="center")

        cols = list(df.columns)
        col_estado = cols.index("Estado") + 1 if "Estado" in cols else None

        for cn, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=cn)
            cell.value = col; cell.fill = navy
            cell.font  = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = c; cell.border = brd

        for rn in range(2, len(df) + 2):
            fondo = blanco if rn % 2 == 0 else gris
            for cn in range(1, len(cols) + 1):
                cell = ws.cell(row=rn, column=cn)
                cell.border = brd; cell.fill = fondo
                cell.alignment = c if cn == col_estado else iz

            if col_estado:
                ce  = ws.cell(row=rn, column=col_estado)
                val = str(ce.value or "").upper()
                if "DESCARGADO" in val:
                    ce.fill = verde; ce.font = Font(bold=True, color="FFFFFF", size=10)
                elif "NO_ENCONTRADO" in val or "ERROR" in val or "TIMEOUT" in val:
                    ce.fill = rojo;  ce.font = Font(bold=True, color="FFFFFF", size=10)
                elif "PROCESANDO" in val:
                    ce.fill = amber; ce.font = Font(bold=True, color="FFFFFF", size=10)

        anchos = {"#":5,"Tipo Doc":10,"Nº Documento":16,"Amparo":12,"Nombre Víctima":30,
                  "Póliza":20,"Siniestro":14,"Placa":10,"Estado":16,"Archivo PDF":22,"Error":35}
        for cn, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(cn)].width = anchos.get(col, 15)
        ws.row_dimensions[1].height = 28
        for rn in range(2, len(df) + 2):
            ws.row_dimensions[rn].height = 17
        ws.freeze_panes = "A2"

    buf.seek(0)
    return buf.read()


# ─── RUTAS FLASK ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    print("👉 Cargando template desde:", os.getcwd())
    return render_template("index.html")


@app.route("/iniciar", methods=["POST"])
def iniciar():
    if estado["corriendo"]:
        return jsonify({"error": "Ya hay un proceso en curso"}), 400

    f = request.files.get("excel")
    if not f:
        return jsonify({"error": "No se adjuntó el archivo Excel"}), 400

    try:
        df = pd.read_excel(f, dtype=str)
        df.columns = df.columns.str.strip().str.upper()
        required = {"TIPO", "ID", "TIPO DE AMPARO"}
        missing  = required - set(df.columns)
        if missing:
            return jsonify({"error": f"Faltan columnas: {', '.join(missing)}"}), 400
        df = df.dropna(subset=["ID"])
        df["TIPO"]           = df["TIPO"].str.strip().str.upper()
        df["ID"]             = df["ID"].str.strip()
        df["TIPO DE AMPARO"] = df["TIPO DE AMPARO"].str.strip().str.upper()
        registros = df.to_dict("records")
    except Exception as e:
        return jsonify({"error": f"Error leyendo Excel: {e}"}), 400

    cfg = {
        "usuario":   request.form.get("usuario",   "PREV900002780"),
        "password":  request.form.get("password",  ""),
        "login_url": request.form.get("login_url", "https://activa-it.net/Login.aspx"),
        "delay":     float(request.form.get("delay", 2)),
        "headless":  request.form.get("headless", "false").lower() == "true",
    }

    # Reset estado
    estado["corriendo"]  = True
    estado["terminado"]  = False
    estado["progreso"]   = []
    estado["pdfs"]       = {}
    estado["total"]      = len(registros)
    estado["config"]     = cfg
    while not estado["log_queue"].empty():
        estado["log_queue"].get_nowait()

    hilo = threading.Thread(target=hilo_playwright, args=(cfg, registros), daemon=True)
    hilo.start()

    return jsonify({"ok": True, "total": len(registros)})


@app.route("/stream")
def stream():
    """Server-Sent Events para progreso en tiempo real."""
    def generate():
        yield "data: {\"tipo\":\"conectado\"}\n\n"
        while True:
            try:
                msg = estado["log_queue"].get(timeout=30)
                yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                if msg.get("tipo") == "fin":
                    break
            except queue.Empty:
                yield "data: {\"tipo\":\"ping\"}\n\n"
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/estado")
def get_estado():
    return jsonify({
        "corriendo":  estado["corriendo"],
        "terminado":  estado["terminado"],
        "total":      estado["total"],
        "progreso":   len(estado["progreso"]),
        "descargados":sum(1 for r in estado["progreso"] if r.get("estado") == "DESCARGADO"),
    })


@app.route("/descargar-excel", methods=["POST"])
def descargar_excel():
    if not estado["progreso"]:
        return jsonify({"error": "Sin datos"}), 400
    excel_bytes = generar_excel(estado["progreso"])
    nombre = f"reporte_descargas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(io.BytesIO(excel_bytes),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=nombre)


@app.route("/descargar-zip", methods=["POST"])
def descargar_zip():
    import zipfile
    if not estado["progreso"]:
        return jsonify({"error": "Sin datos"}), 400

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for nombre, pdf_bytes in estado["pdfs"].items():
            zf.writestr(f"certificados/{nombre}", pdf_bytes)
        excel_bytes = generar_excel(estado["progreso"])
        zf.writestr(f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", excel_bytes)

    buf.seek(0)
    zip_nombre = f"certificados_tope_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return send_file(buf, mimetype="application/zip",
                     as_attachment=True, download_name=zip_nombre)

@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


if __name__ == "__main__":
    app.run(debug=False, host="127.0.0.1", port=5001)
